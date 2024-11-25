import datetime
import random
from openpyxl import Workbook, load_workbook

# 数据存储
goods_dict = {}  # 商品信息：条码 -> {"name": 品名, "price": 单价, "quantity": 数量}
sales_data = {}  # 销售统计：条码 -> {"name": 品名, "sales": 销量}
total_sales = 0  # 总销售额


def save_goods():
    """保存商品信息到Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "商品信息"
    ws.append(["条形码", "商品名称", "单价", "数量"])

    for bar_code, info in goods_dict.items():
        ws.append([bar_code, info['name'], info['price'], info['quantity']])

    # 保存销售统计信息到另一个工作表
    wb.create_sheet(title="销售统计")
    save_statistics(wb)
    wb.save("商品与销售统计.xlsx")
    print("商品信息已保存到Excel文件。")


def load_goods():
    """从Excel文件加载商品信息"""
    try:
        wb = load_workbook("商品与销售统计.xlsx")
        ws = wb["商品信息"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            bar_code, name, price, quantity = row
            goods_dict[bar_code] = {
                "name": name,
                "price": price,
                "quantity": quantity,
            }
            sales_data[bar_code] = {"name": name, "sales": 0}
    except FileNotFoundError:
        print("商品信息文件未找到，创建新文件。")


def load_statistics():
    """从销售统计工作表加载已有的统计数据"""
    global total_sales
    try:
        wb = load_workbook("商品与销售统计.xlsx")
        ws = wb["销售统计"]
        total_sales = ws.cell(row=2, column=2).value
        print(f"最新销售时间: {ws.cell(row=1, column=2).value}")
    except FileNotFoundError:
        print("销售统计文件未找到，创建新文件。")


def update_total_sales():
    """更新总销售额"""
    global total_sales
    total_sales = 0  # 先初始化为0，后续进行累加
    for bar_code, info in sales_data.items():
        item_sales = info["sales"] * goods_dict[bar_code]["price"]
        total_sales += item_sales


def save_statistics(wb=None):
    """保存销售统计信息到Excel文件的销售统计工作表"""
    global total_sales
    if wb is None:
        wb = load_workbook("商品与销售统计.xlsx")
    ws = wb["销售统计"]

    # 更新最新销售时间和总销售额
    ws.cell(row=1, column=1, value="最新销售时间")
    ws.cell(row=1, column=2, value=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    ws.cell(row=2, column=1, value="总销售额")
    ws.cell(row=2, column=2, value=total_sales)

    # 排序销量前几的商品
    sorted_sales = sorted(sales_data.items(), key=lambda x: x[1]["sales"], reverse=True)
    ws.cell(row=4, column=1, value="热销商品前三：")
    row_num = 5
    for idx, (bar_code, info) in enumerate(sorted_sales[:3], start=1):
        item_sales = info["sales"] * goods_dict[bar_code]["price"]
        ws.cell(row=row_num, column=1, value=f"第{idx}名：{info['name']}")
        ws.cell(row=row_num, column=2, value=f"销量：{info['sales']}")
        ws.cell(row=row_num, column=3, value=f"销售额：{item_sales:.2f}")
        row_num += 1

    ws.cell(row=row_num, column=1, value="其他商品销量：")
    row_num += 1
    for bar_code, info in sorted_sales[3:]:
        item_sales = info["sales"] * goods_dict[bar_code]["price"]
        ws.cell(row=row_num, column=1, value=info['name'])
        ws.cell(row=row_num, column=2, value=f"销量：{info['sales']}")
        ws.cell(row=row_num, column=3, value=f"销售额：{item_sales:.2f}")
        row_num += 1

    wb.save("商品与销售统计.xlsx")
    print("销售统计信息已保存到Excel文件。")




def add_goods():
    """增加商品"""
    while True:
        bar_code = input("请输入商品条码号 (按回车退出): ")
        if bar_code == "":  # 如果输入为空，退出循环
            break
        if bar_code in goods_dict:
            goods_dict[bar_code]["quantity"] += 1  # 自动加一
            print(f"商品 {goods_dict[bar_code]['name']} 已存在，数量+1，当前数量为 {goods_dict[bar_code]['quantity']}。\n")
        else:
            name = input("请输入商品名称: ")
            while True:
                try:
                    price = input("请输入商品单价: ")
                    price = float(price)
                    break
                except ValueError:
                    print("无效价格，请重试。\n")

            quantity = input("请输入商品数量: ")
            if not quantity.isdigit():
                print("无效数量，请重试。\n")
                continue
            goods_dict[bar_code] = {
                "name": name,
                "price": float(price),
                "quantity": int(quantity),
            }
            sales_data[bar_code] = {"name": name, "sales": 0}
            print(f"新商品 {name} 已添加！\n")

        save_goods()
        save_statistics()

    checkout_choice = input("是否立即进行收银？(y/n): ").strip().lower()
    if checkout_choice == 'y':
        checkout()
    elif checkout_choice == 'n':
        print("返回主菜单。\n")
    else:
        print("无效输入，请选择 'y' 或 'n'。\n")


def delete_goods():
    """删除商品"""
    bar_code = input("请输入要删除的商品条码号: ")
    if bar_code in goods_dict:
        del goods_dict[bar_code]
        del sales_data[bar_code]
        print("商品已删除！\n")
        save_goods()
        save_statistics()
    else:
        print("未找到该商品。\n")


def update_goods():
    """修改商品"""
    bar_code = input("请输入商品条码号: ")
    if bar_code in goods_dict:
        print(f"当前商品信息：{goods_dict[bar_code]}")
        name = input("新的商品名称 (回车跳过): ") or goods_dict[bar_code]["name"]
        price = input("新的单价 (回车跳过): ")
        price = float(price) if price else goods_dict[bar_code]["price"]
        quantity = input("新的数量数量 (回车跳过): ")
        quantity = int(quantity) if quantity else goods_dict[bar_code]["quantity"]
        goods_dict[bar_code] = {"name": name, "price": price, "quantity": quantity}
        print("商品信息已更新！\n")
        save_goods()
        save_statistics()
    else:
        print("未找到该商品。\n")


def query_goods():
    """查询商品"""
    bar_code = input("请输入商品条码号: ")
    if bar_code in goods_dict:
        info = goods_dict[bar_code]
        print(f"商品信息：\n条码号:{bar_code}\n名称:{info['name']}\n单价:{info['price']}\n数量:{info['quantity']}\n")
    else:
        print("未找到该商品。\n")


def statistics():
    """显示统计信息"""
    if total_sales == 0:
        print("\n\n尚无任何销售记录，请先进行收银。")
    print(f"交易总销售额：{total_sales:.2f}元")
    sorted_sales = sorted(sales_data.items(), key=lambda x: x[1]["sales"], reverse=True)
    print("热销商品前三：")
    for idx, (bar_code, info) in enumerate(sorted_sales[:3], start=1):
        print(f"{idx}. {info['name']} - 销量：{info['sales']}")
    print("\n其他商品销量：")
    for bar_code, info in sorted_sales[3:]:
        print(f"{info['name']} - 销量：{info['sales']}")


def print_receipt(items, pricesinglelist, itemnumbers, pricelist, sumup, sumnum):
    """打印小票"""
    ticket = random.randint(211111111111111111, 299999999999999999)
    frontdesk = random.randint(1, 99)
    casher = random.randint(1000000000, 9999999999)

    print("\n-------------------------------------------")
    print("青青草原超市（欢迎光临）")
    print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    print(f"票单：{ticket}")
    print(f"款台: {frontdesk}  收银员: {casher}")
    print("-------------------------------------------")
    print(f"{'商品名':<12} {'单价':<10} {'数量':<8} {'总价':<10}")
    for i in range(len(items)):
        print(f"{items[i]:<12} {pricesinglelist[i]:<10.2f} {itemnumbers[i]:<8} {pricelist[i]:<10.2f}")
    print("-------------------------------------------")
    print(f"金额总计：{sumup:.2f} 元")
    print(f"数量总计：{sumnum} 件")
    print("-------------------------------------------")
    print("谢谢惠顾！期待你的下次光临！")


def checkout():
    """收银"""
    global total_sales
    items = []
    pricesinglelist = []
    itemnumbers = []
    pricelist = []
    total_price = 0
    total_num = 0

    # 循环遍历商品，并使用已经存储的数量
    for bar_code, info in goods_dict.items():
        quantity = info['quantity']  # 使用添加商品时输入的数量
        if quantity <= 0:
            print(f"商品 {info['name']} 数量不足，跳过该商品。\n")
            continue
        price = info['price'] * quantity
        items.append(info['name'])
        pricesinglelist.append(info['price'])
        itemnumbers.append(quantity)
        pricelist.append(price)
        total_price += price
        total_num += quantity
        sales_data[bar_code]["sales"] += quantity  # 更新销量

    if total_num > 0:
        print_receipt(items, pricesinglelist, itemnumbers, pricelist, total_price, total_num)

        # 付款码验证
        while True:
            cod = input("请出示付款码 (18位，前两位为10-15): ")
            if not cod:  # 如果用户输入空的收款码，取消支付
                print("支付已取消。\n")
                return  # 取消支付后退出收银
            if len(cod) == 18 and cod[:2] in {"10", "11", "12", "13", "14", "15"}:
                print("付款成功！\n")
                break
            else:
                print("不是有效付款码，请重试。\n")

        update_total_sales()  # 更新总销售额
        save_goods()  # 保存商品信息
        save_statistics()  # 保存销售统计信息
    else:
        print("没有选择任何商品，收银取消。\n")




def main():
    """主菜单"""
    load_goods()
    load_statistics()

    while True:
        print("\n\n")
        print("1. 增加商品")
        print("2. 删除商品")
        print("3. 修改商品")
        print("4. 查询商品")
        print("5. 统计信息")
        print("6. 收银")
        print("7. 退出系统")

        choice = input("请输入选择：")
        if choice == "1":
            add_goods()
        elif choice == "2":
            delete_goods()
        elif choice == "3":
            update_goods()
        elif choice == "4":
            query_goods()
        elif choice == "5":
            statistics()
        elif choice == "6":
            checkout()
        elif choice == "7":
            print("感谢使用，系统已退出。")
            break
        else:
            print("无效输入，请重新选择。\n")


if __name__ == "__main__":
    main()